import json
import os
import shutil
import tempfile

from django.test import SimpleTestCase, TestCase
from openpyxl import load_workbook

from karate_tests import runner
from karate_tests.models import KarateTestCaseJob

FIXTURES_DIR = os.path.join(os.path.dirname(__file__), 'fixtures')
SAMPLE_REPORT = os.path.join(FIXTURES_DIR, 'simple_feature_report.html')
OVERVIEW_REPORT = os.path.join(FIXTURES_DIR, 'overview_no_data.html')
MALFORMED_REPORT = os.path.join(FIXTURES_DIR, 'malformed_data.html')
CALLER_REPORT = os.path.join(FIXTURES_DIR, 'caller_feature_report.html')


class ValidateReportsDirTests(SimpleTestCase):
    def test_rejects_empty(self):
        with self.assertRaises(runner.PreflightError):
            runner.validate_reports_dir('')

    def test_rejects_relative_path(self):
        with self.assertRaises(runner.PreflightError):
            runner.validate_reports_dir('relative/path')

    def test_rejects_nonexistent_directory(self):
        with self.assertRaises(runner.PreflightError):
            runner.validate_reports_dir('/tmp/qa-helper-tool-does-not-exist-xyz')

    def test_accepts_existing_directory(self):
        with tempfile.TemporaryDirectory() as d:
            self.assertEqual(runner.validate_reports_dir(d), d)


class ValidateExcelPathTests(SimpleTestCase):
    def test_rejects_empty(self):
        with self.assertRaises(runner.PreflightError):
            runner.validate_excel_path('')

    def test_rejects_relative_path(self):
        with self.assertRaises(runner.PreflightError):
            runner.validate_excel_path('relative/output.xlsx')

    def test_rejects_non_xlsx_extension(self):
        with self.assertRaises(runner.PreflightError):
            runner.validate_excel_path('/tmp/output.csv')

    def test_accepts_absolute_xlsx_path(self):
        path = '/tmp/qa-helper-tool-output.xlsx'
        self.assertEqual(runner.validate_excel_path(path), path)


class ParseReportFileTests(SimpleTestCase):
    def test_extracts_feature_json_from_real_sample(self):
        feature = runner.parse_report_file(SAMPLE_REPORT)
        self.assertIsNotNone(feature)
        self.assertEqual(feature['name'], 'simple crud flow against the todo API')
        self.assertEqual(len(feature['scenarios']), 1)
        self.assertEqual(feature['scenarios'][0]['name'], 'simple crud flow')

    def test_returns_none_for_page_without_karate_data(self):
        self.assertIsNone(runner.parse_report_file(OVERVIEW_REPORT))

    def test_raises_for_malformed_json(self):
        with self.assertRaises(json.JSONDecodeError):
            runner.parse_report_file(MALFORMED_REPORT)


class FindReportFilesTests(SimpleTestCase):
    def test_finds_html_files_recursively(self):
        with tempfile.TemporaryDirectory() as d:
            nested = os.path.join(d, 'nested')
            os.makedirs(nested)
            shutil.copy(SAMPLE_REPORT, os.path.join(d, 'a.html'))
            shutil.copy(SAMPLE_REPORT, os.path.join(nested, 'b.html'))
            with open(os.path.join(d, 'ignored.txt'), 'w') as f:
                f.write('not html')
            found = runner.find_report_files(d)
            self.assertEqual(len(found), 2)
            self.assertTrue(all(p.endswith('.html') for p in found))


class ExtractApiStepsTests(SimpleTestCase):
    def setUp(self):
        self.feature = runner.parse_report_file(SAMPLE_REPORT)
        self.scenario = self.feature['scenarios'][0]

    def test_extracts_one_step_per_http_call_in_order(self):
        steps = runner.extract_api_steps(self.scenario)
        self.assertEqual([s['method'] for s in steps], ['POST', 'GET', 'GET', 'POST', 'GET'])

    def test_post_step_has_status_and_response_body(self):
        steps = runner.extract_api_steps(self.scenario)
        first = steps[0]
        self.assertEqual(first['status_code'], '201')
        self.assertIn('"title": "First"', first['response_body'])
        self.assertIn('/api/todos', first['url'])

    def test_post_step_curl_includes_method_url_and_body(self):
        steps = runner.extract_api_steps(self.scenario)
        curl = steps[0]['curl']
        self.assertIn('curl -X POST', curl)
        self.assertIn('/api/todos', curl)
        self.assertIn("-d '", curl)
        self.assertIn('"title": "First"', curl)

    def test_post_step_curl_excludes_noisy_headers(self):
        steps = runner.extract_api_steps(self.scenario)
        curl = steps[0]['curl']
        self.assertNotIn('Content-Length', curl)
        self.assertNotIn('Host:', curl)
        self.assertNotIn('User-Agent', curl)
        # Content-Type is meaningful and should survive.
        self.assertIn('Content-Type', curl)

    def test_get_step_has_no_request_body_in_curl(self):
        steps = runner.extract_api_steps(self.scenario)
        get_step = steps[1]
        self.assertEqual(get_step['method'], 'GET')
        self.assertNotIn('-d ', get_step['curl'])
        self.assertEqual(get_step['status_code'], '200')


class ExtractApiStepsSkippedStepTests(SimpleTestCase):
    """A 'method' step that Karate marks status='skipped' never actually
    ran (an earlier step in the scenario failed first) -- there's genuinely
    no request to replay, so it's dropped rather than shown as a phantom
    "Execute the CURL" row."""

    def test_skipped_step_is_omitted(self):
        scenario = {
            'name': 'a scenario',
            'steps': [{'keyword': 'method', 'text': 'post', 'line': 10, 'status': 'skipped', 'logSegments': []}],
        }
        self.assertEqual(runner.extract_api_steps(scenario), [])

    def test_skipped_step_produces_a_warning(self):
        scenario = {
            'name': 'a scenario',
            'steps': [{'keyword': 'method', 'text': 'post', 'line': 10, 'status': 'skipped', 'logSegments': []}],
        }
        warnings = []
        runner.extract_api_steps(scenario, warnings=warnings)
        self.assertEqual(len(warnings), 1)
        self.assertIn('a scenario', warnings[0])
        self.assertIn('post', warnings[0])
        self.assertIn('line 10', warnings[0])
        self.assertIn('skipped', warnings[0])

    def test_other_steps_in_the_same_scenario_are_unaffected(self):
        scenario = {
            'name': 'a scenario',
            'steps': [
                {'keyword': 'method', 'text': 'post', 'line': 10, 'status': 'skipped', 'logSegments': []},
                {
                    'keyword': 'method', 'text': 'get', 'line': 15, 'status': 'passed',
                    'logSegments': [{'text': '1 > GET http://x\n\nresponse time in milliseconds: 1\n1 < 200 GET http://x\n'}],
                },
            ],
        }
        steps = runner.extract_api_steps(scenario)
        self.assertEqual(len(steps), 1)
        self.assertEqual(steps[0]['method'], 'GET')


class ExtractApiStepsUnparsableStepTests(SimpleTestCase):
    """A step that actually executed (status isn't 'skipped') can still end
    up with no usable log data -- e.g. some Karate configs/versions omit
    detailed request/response logging even for calls that ran, which in
    practice hit non-GET methods harder than bodyless GETs. Unlike a
    genuinely skipped step, this one really did make a call, so it's kept
    (with the existing "could not reconstruct" placeholder) rather than
    dropped -- dropping these caused real executed POST/PUT/DELETE calls to
    silently vanish while GET calls kept working, and whole reports to fail
    once every step in them was affected this way."""

    def test_step_with_no_log_segments_and_passed_status_still_produces_a_row(self):
        scenario = {
            'name': 'a scenario',
            'steps': [{'keyword': 'method', 'text': 'post', 'line': 10, 'status': 'passed', 'logSegments': []}],
        }
        steps = runner.extract_api_steps(scenario)
        self.assertEqual(len(steps), 1)
        self.assertIn('# Could not reconstruct the request for step: post', steps[0]['curl'])

    def test_step_with_no_status_field_still_produces_a_row(self):
        # Older/other Karate report variants may not include a 'status'
        # field at all -- absence must not be treated as 'skipped'.
        scenario = {
            'name': 'a scenario',
            'steps': [{'keyword': 'method', 'text': 'post', 'line': 10}],
        }
        steps = runner.extract_api_steps(scenario)
        self.assertEqual(len(steps), 1)

    def test_step_with_no_logs_produces_a_warning(self):
        scenario = {
            'name': 'a scenario',
            'steps': [{'keyword': 'method', 'text': 'post', 'line': 10, 'status': 'passed', 'logSegments': []}],
        }
        warnings = []
        runner.extract_api_steps(scenario, warnings=warnings)
        self.assertEqual(len(warnings), 1)
        self.assertIn('a scenario', warnings[0])
        self.assertIn('post', warnings[0])
        self.assertIn('line 10', warnings[0])

    def test_step_with_logs_but_no_request_line_falls_back_with_a_warning(self):
        # logSegments present, but the text doesn't contain a recognizable
        # 'N > METHOD URL' request line -- a genuine parsing gap.
        scenario = {
            'name': 'a scenario',
            'steps': [{
                'keyword': 'method', 'text': 'post', 'line': 12, 'status': 'passed',
                'logSegments': [{'text': 'some unexpected log format with no request line'}],
            }],
        }
        warnings = []
        steps = runner.extract_api_steps(scenario, warnings=warnings)
        self.assertEqual(len(steps), 1)
        self.assertIn('# Could not reconstruct the request for step: post', steps[0]['curl'])
        self.assertEqual(len(warnings), 1)
        self.assertIn('could not be reconstructed', warnings[0])


class ExtractApiStepsWithCallResultsTests(SimpleTestCase):
    """Real Karate report where the scenario's first step is
    `created = call read('called.feature') {...}` -- the called feature
    makes its own API call before the caller makes its own."""

    def setUp(self):
        self.feature = runner.parse_report_file(CALLER_REPORT)
        self.scenario = self.feature['scenarios'][0]

    def test_includes_calls_made_by_the_called_feature(self):
        steps = runner.extract_api_steps(self.scenario)
        # POST happens inside called.feature; GET is the caller's own step.
        self.assertEqual([s['method'] for s in steps], ['POST', 'GET'])

    def test_order_matches_actual_execution_order(self):
        steps = runner.extract_api_steps(self.scenario)
        self.assertIn('/api/todos', steps[0]['url'])
        self.assertNotIn('/api/todos/', steps[0]['url'])  # the create call, no id in path
        self.assertIn('/api/todos/', steps[1]['url'])  # the caller's own fetch-by-id

    def test_called_feature_step_has_correct_data(self):
        steps = runner.extract_api_steps(self.scenario)
        created = steps[0]
        self.assertEqual(created['status_code'], '201')
        self.assertIn('"title": "from-caller"', created['response_body'])
        self.assertIn('curl -X POST', created['curl'])

    def test_caller_own_step_has_correct_data(self):
        steps = runner.extract_api_steps(self.scenario)
        fetched = steps[1]
        self.assertEqual(fetched['status_code'], '200')
        self.assertIn('"title": "from-caller"', fetched['response_body'])

    def test_build_test_cases_includes_both_steps(self):
        cases = runner.build_test_cases(self.feature)
        self.assertEqual(len(cases), 1)
        self.assertEqual(len(cases[0]['steps']), 2)


class BuildTestCasesTests(SimpleTestCase):
    def test_builds_one_case_per_scenario_named_with_scenario_name_only(self):
        feature = runner.parse_report_file(SAMPLE_REPORT)
        cases = runner.build_test_cases(feature)
        self.assertEqual(len(cases), 1)
        self.assertEqual(cases[0]['name'], 'simple crud flow')
        self.assertEqual(len(cases[0]['steps']), 5)

    def test_falls_back_to_feature_name_when_scenario_has_no_name(self):
        feature = {
            'name': 'fallback feature',
            'scenarios': [{
                'name': '',
                'steps': [{
                    'keyword': 'method', 'text': 'get',
                    'logSegments': [{'text': '1 > GET http://x\n\nresponse time in milliseconds: 1\n1 < 200 GET http://x\n'}],
                }],
            }],
        }
        cases = runner.build_test_cases(feature)
        self.assertEqual(cases[0]['name'], 'fallback feature')

    def test_scenario_with_no_http_calls_is_skipped(self):
        feature = {
            'name': 'no-op feature',
            'scenarios': [{'name': 'just some math', 'steps': [{'keyword': 'def', 'text': 'x = 1'}]}],
        }
        self.assertEqual(runner.build_test_cases(feature), [])


class BuildCurlTests(SimpleTestCase):
    def test_returns_empty_string_without_method_or_url(self):
        self.assertEqual(runner.build_curl(None, None, {}, None), '')

    def test_quotes_values_containing_single_quotes(self):
        curl = runner.build_curl('GET', "http://example.com/it's", {}, None)
        self.assertIn('"http://example.com/it\'s"', curl)


class GenerateTests(TestCase):
    def _make_job(self, reports_dir, excel_path, **overrides):
        defaults = dict(
            reports_dir=reports_dir,
            excel_path=excel_path,
            environment='QA',
            pre_requisite='User is logged in',
            created_by='Ada',
            sprint='Sprint 1',
            status=KarateTestCaseJob.STATUS_RUNNING,
        )
        defaults.update(overrides)
        return KarateTestCaseJob.objects.create(**defaults)

    def test_end_to_end_produces_expected_workbook(self):
        with tempfile.TemporaryDirectory() as reports_dir, tempfile.TemporaryDirectory() as out_dir:
            shutil.copy(SAMPLE_REPORT, os.path.join(reports_dir, 'simple.html'))
            excel_path = os.path.join(out_dir, 'cases.xlsx')
            job = self._make_job(reports_dir, excel_path)

            runner.generate(job)
            job.refresh_from_db()

            self.assertEqual(job.status, KarateTestCaseJob.STATUS_COMPLETED)
            self.assertEqual(job.feature_count, 1)
            self.assertEqual(job.scenario_count, 1)
            self.assertEqual(job.step_count, 5)
            self.assertTrue(os.path.isfile(excel_path))

            wb = load_workbook(excel_path)
            ws = wb.active
            header = [c.value for c in ws[1]]
            self.assertEqual(header, runner.COLUMNS)

            rows = list(ws.iter_rows(min_row=2, values_only=True))
            self.assertEqual(len(rows), 5)

            first = rows[0]
            self.assertEqual(first[0], 1)  # S.No
            self.assertEqual(first[1], 'simple crud flow')
            self.assertEqual(first[2], first[1])  # description == name
            self.assertEqual(first[3], 'QA')
            self.assertEqual(first[4], 'User is logged in')
            self.assertEqual(first[5], 1)
            self.assertEqual(first[6], 'Execute the CURL')
            self.assertIn('curl -X POST', first[7])
            self.assertIn('Response Code: 201', first[8])
            self.assertIn('Response Code: 201', first[9])
            self.assertEqual(first[10], 'Ada')
            self.assertEqual(first[11], 'Sprint 1')

            # Step numbers increment per test case.
            self.assertEqual([r[5] for r in rows], [1, 2, 3, 4, 5])

            # Case-level columns are blank on every row after the first --
            # the value lives only in the top-left cell of the merged range.
            for row in rows[1:]:
                self.assertIsNone(row[0])
                self.assertIsNone(row[1])
                self.assertIsNone(row[2])
                self.assertIsNone(row[3])
                self.assertIsNone(row[4])
                self.assertIsNone(row[10])
                self.assertIsNone(row[11])

            merged_ranges = {str(r) for r in ws.merged_cells.ranges}
            for col_letter in ('A', 'B', 'C', 'D', 'E', 'K', 'L'):
                self.assertIn(f'{col_letter}2:{col_letter}6', merged_ranges)

    def test_end_to_end_includes_steps_from_called_feature(self):
        with tempfile.TemporaryDirectory() as reports_dir, tempfile.TemporaryDirectory() as out_dir:
            shutil.copy(CALLER_REPORT, os.path.join(reports_dir, 'caller.html'))
            excel_path = os.path.join(out_dir, 'cases.xlsx')
            job = self._make_job(reports_dir, excel_path)

            runner.generate(job)
            job.refresh_from_db()

            self.assertEqual(job.status, KarateTestCaseJob.STATUS_COMPLETED)
            self.assertEqual(job.scenario_count, 1)
            self.assertEqual(job.step_count, 2)

            wb = load_workbook(excel_path)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            self.assertEqual(len(rows), 2)
            self.assertIn('curl -X POST', rows[0][7])
            self.assertIn('curl -X GET', rows[1][7])

    def test_single_step_case_has_no_merge(self):
        with tempfile.TemporaryDirectory() as reports_dir, tempfile.TemporaryDirectory() as out_dir:
            feature_html = SAMPLE_REPORT
            # Build a single-call feature by trimming the sample's scenario
            # steps down to just the first HTTP call, to exercise the
            # no-merge-needed path (a case with only one step).
            content = open(feature_html, encoding='utf-8').read()
            match = runner._KARATE_DATA_RE.search(content)
            feature = json.loads(match.group(1))
            scenario = feature['scenarios'][0]
            method_seen = False
            trimmed_steps = []
            for step in scenario['steps']:
                trimmed_steps.append(step)
                if step.get('keyword') == 'method':
                    method_seen = True
                    break
            self.assertTrue(method_seen)
            scenario['steps'] = trimmed_steps
            feature['scenarios'] = [scenario]

            single_call_html = (
                '<script id="karate-data" type="application/json">'
                + json.dumps(feature)
                + '</script>'
            )
            path = os.path.join(reports_dir, 'single.html')
            with open(path, 'w', encoding='utf-8') as f:
                f.write(single_call_html)

            excel_path = os.path.join(out_dir, 'cases.xlsx')
            job = self._make_job(reports_dir, excel_path)
            runner.generate(job)
            job.refresh_from_db()

            self.assertEqual(job.status, KarateTestCaseJob.STATUS_COMPLETED)
            self.assertEqual(job.step_count, 1)

            wb = load_workbook(excel_path)
            ws = wb.active
            self.assertEqual(len(ws.merged_cells.ranges), 0)
            row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
            self.assertIsNotNone(row[0])  # value present directly, no merge needed

    def test_no_html_files_marks_failed(self):
        with tempfile.TemporaryDirectory() as reports_dir, tempfile.TemporaryDirectory() as out_dir:
            excel_path = os.path.join(out_dir, 'cases.xlsx')
            job = self._make_job(reports_dir, excel_path)
            runner.generate(job)
            job.refresh_from_db()
            self.assertEqual(job.status, KarateTestCaseJob.STATUS_FAILED)
            self.assertIn('No Karate feature HTML reports found', job.error)
            self.assertFalse(os.path.isfile(excel_path))

    def test_reports_with_no_http_calls_marks_failed(self):
        with tempfile.TemporaryDirectory() as reports_dir, tempfile.TemporaryDirectory() as out_dir:
            shutil.copy(OVERVIEW_REPORT, os.path.join(reports_dir, 'overview.html'))
            excel_path = os.path.join(out_dir, 'cases.xlsx')
            job = self._make_job(reports_dir, excel_path)
            runner.generate(job)
            job.refresh_from_db()
            self.assertEqual(job.status, KarateTestCaseJob.STATUS_FAILED)

    def test_executed_steps_without_logs_still_produce_a_completed_job(self):
        # Regression: steps that genuinely ran (status='passed') but have no
        # logSegments -- e.g. a Karate config/version that doesn't attach
        # detailed request/response logging -- must not be dropped. Dropping
        # them previously caused a report where *every* step was like this
        # to end up with zero cases and a spurious "no scenario made any
        # HTTP calls" failure, even though calls were actually made -- and
        # in reports with a mix of methods, it disproportionately hid
        # non-GET calls (bodyless GETs were more likely to still log
        # cleanly), making it look like only GET requests were recognized.
        with tempfile.TemporaryDirectory() as reports_dir, tempfile.TemporaryDirectory() as out_dir:
            feature = {
                'name': 'a feature',
                'scenarios': [{
                    'name': 'a scenario',
                    'steps': [
                        {'keyword': 'method', 'text': 'post', 'line': 10, 'status': 'passed', 'logSegments': []},
                        {'keyword': 'method', 'text': 'get', 'line': 15, 'status': 'passed', 'logSegments': []},
                    ],
                }],
            }
            html = '<script id="karate-data" type="application/json">' + json.dumps(feature) + '</script>'
            path = os.path.join(reports_dir, 'no_logs.html')
            with open(path, 'w', encoding='utf-8') as f:
                f.write(html)

            excel_path = os.path.join(out_dir, 'cases.xlsx')
            job = self._make_job(reports_dir, excel_path)
            runner.generate(job)
            job.refresh_from_db()

            self.assertEqual(job.status, KarateTestCaseJob.STATUS_COMPLETED)
            self.assertEqual(job.step_count, 2)
            self.assertEqual(len(job.warnings), 2)

            wb = load_workbook(excel_path)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            self.assertEqual(len(rows), 2)
            self.assertIn('Could not reconstruct', rows[0][7])
            self.assertIn('Could not reconstruct', rows[1][7])

    def test_skipped_step_is_omitted_and_reported_as_a_warning(self):
        with tempfile.TemporaryDirectory() as reports_dir, tempfile.TemporaryDirectory() as out_dir:
            feature = {
                'name': 'a feature',
                'scenarios': [{
                    'name': 'a scenario',
                    'steps': [
                        {'keyword': 'method', 'text': 'post', 'line': 10, 'status': 'skipped', 'logSegments': []},
                        {
                            'keyword': 'method', 'text': 'get', 'line': 15, 'status': 'passed',
                            'logSegments': [{'text': '1 > GET http://x\n\nresponse time in milliseconds: 1\n1 < 200 GET http://x\n'}],
                        },
                    ],
                }],
            }
            html = '<script id="karate-data" type="application/json">' + json.dumps(feature) + '</script>'
            path = os.path.join(reports_dir, 'partial.html')
            with open(path, 'w', encoding='utf-8') as f:
                f.write(html)

            excel_path = os.path.join(out_dir, 'cases.xlsx')
            job = self._make_job(reports_dir, excel_path)
            runner.generate(job)
            job.refresh_from_db()

            self.assertEqual(job.status, KarateTestCaseJob.STATUS_COMPLETED)
            self.assertEqual(job.step_count, 1)  # the skipped 'post' step is omitted
            self.assertEqual(len(job.warnings), 1)
            self.assertIn('partial.html', job.warnings[0])
            self.assertIn('post', job.warnings[0])

    def test_malformed_report_is_collected_as_warning_not_fatal(self):
        with tempfile.TemporaryDirectory() as reports_dir, tempfile.TemporaryDirectory() as out_dir:
            shutil.copy(SAMPLE_REPORT, os.path.join(reports_dir, 'good.html'))
            shutil.copy(MALFORMED_REPORT, os.path.join(reports_dir, 'bad.html'))
            excel_path = os.path.join(out_dir, 'cases.xlsx')
            job = self._make_job(reports_dir, excel_path)
            runner.generate(job)
            job.refresh_from_db()
            self.assertEqual(job.status, KarateTestCaseJob.STATUS_COMPLETED)
            self.assertEqual(len(job.warnings), 1)
            self.assertIn('bad.html', job.warnings[0])
