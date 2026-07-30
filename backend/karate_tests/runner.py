"""Turns Karate HTML feature reports into an Excel test-case sheet.

Modern Karate reports (the Alpine.js-based template) embed each feature's
full execution data as JSON in a `<script id="karate-data">` tag -- the
actual HTML is just a client-side rendering shell. We read that JSON
directly rather than scraping rendered markup.

For each scenario, every step with keyword 'method' is the point where
Karate actually fires an HTTP call (preceding steps like 'url'/'request'/
'header' just build it up, following steps like 'status'/'match' just
assert on the result). Each such step's `logSegments` carry the raw
request/response text Karate logged, which we parse back into method, url,
headers, body, status code, and response body -- from which we can
reconstruct an equivalent curl command.

A step that calls another feature (Karate's `call`/`callSingle`, typically
on a `def` step) carries `hasCallResults`/`callResults` -- each entry in
`callResults` is itself a full feature-shaped dict (its own `scenarios` /
`steps`), since Karate executes it synchronously at that point. We walk
those recursively too, so API calls made inside a called feature show up
as their own steps, inlined in the exact order they actually ran.
"""
import glob
import json
import os
import re

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

_KARATE_DATA_RE = re.compile(
    r'<script id="karate-data" type="application/json">(.*?)</script>', re.DOTALL
)
_REQUEST_LINE_RE = re.compile(r'^(\d+) > (.+)$')
_RESPONSE_LINE_RE = re.compile(r'^(\d+) < (.+)$')

# Transport-level headers curl (or the target server) sets/computes itself --
# replaying them verbatim would be noise at best, wrong at worst (e.g. a
# stale Content-Length if the body is edited before replay).
_NOISY_REQUEST_HEADERS = {'content-length', 'host', 'connection', 'user-agent'}

COLUMNS = [
    'S.No', 'Test Case Name', 'Test Case Description', 'Environment', 'Pre-Requisite',
    'Step #', 'Step Description', 'Test Data', 'Expected Result', 'Actual Result',
    'Created By', 'Sprint',
]


class PreflightError(ValueError):
    """Raised for validation problems caught before we scan/write anything."""


def validate_reports_dir(reports_dir):
    reports_dir = (reports_dir or '').strip()
    if not reports_dir:
        raise PreflightError('Provide the folder containing your Karate HTML reports.')
    if not os.path.isabs(reports_dir):
        raise PreflightError('Reports location must be an absolute path.')
    if not os.path.isdir(reports_dir):
        raise PreflightError(f"'{reports_dir}' is not a directory (or doesn't exist).")
    return reports_dir


def validate_excel_path(excel_path):
    excel_path = (excel_path or '').strip()
    if not excel_path:
        raise PreflightError('Provide where the generated Excel file should be saved.')
    if not os.path.isabs(excel_path):
        raise PreflightError('Excel file location must be an absolute path.')
    if not excel_path.lower().endswith(('.xlsx', '.xlsm')):
        raise PreflightError("Excel file location must end in '.xlsx'.")
    return excel_path


def find_report_files(reports_dir):
    return sorted(glob.glob(os.path.join(reports_dir, '**', '*.html'), recursive=True))


def parse_report_file(path):
    """Returns the embedded feature dict, or None if this .html has no
    karate-data script (e.g. an overview/summary page rather than a
    per-feature report)."""
    with open(path, 'r', encoding='utf-8') as f:
        content = f.read()
    match = _KARATE_DATA_RE.search(content)
    if not match:
        return None
    return json.loads(match.group(1))


def _shell_quote(value):
    value = str(value)
    if "'" not in value:
        return f"'{value}'"
    escaped = value.replace('\\', '\\\\').replace('"', '\\"')
    return f'"{escaped}"'


def _split_request_response(log_segments):
    """Walks a 'method' step's logSegments -- alternating text (request/
    response line + headers) and code (request/response body) segments --
    into (method, url, headers, request_body, status_code, response_body).

    For calls with no request body (GET/DELETE), the request and response
    text arrive concatenated in a single segment; we split on the
    'response time in milliseconds' marker Karate always logs there.
    """
    request_text, response_text = '', ''
    request_body, response_body = None, None
    phase = 'request'

    for seg in log_segments:
        text = seg.get('text')
        code = seg.get('code')
        if text is not None:
            idx = text.find('response time in milliseconds')
            if idx == -1:
                marker = re.search(r'\d+ < \d+ ', text)
                idx = marker.start() if marker else -1
            if idx == -1:
                if phase == 'request':
                    request_text += text
                else:
                    response_text += text
            else:
                request_text += text[:idx]
                response_text += text[idx:]
                phase = 'response'
        elif code is not None:
            if phase == 'request':
                request_body = code
            else:
                response_body = code

    method, url, headers = None, None, {}
    seen_request_line = False
    for line in request_text.splitlines():
        m = _REQUEST_LINE_RE.match(line)
        if not m:
            continue
        rest = m.group(2)
        if not seen_request_line:
            parts = rest.split(' ', 1)
            if len(parts) == 2:
                method, url = parts
            seen_request_line = True
        elif ':' in rest:
            key, _, value = rest.partition(':')
            headers[key.strip()] = value.strip()

    status_code = None
    seen_response_line = False
    for line in response_text.splitlines():
        m = _RESPONSE_LINE_RE.match(line)
        if not m:
            continue
        if not seen_response_line:
            status_code = m.group(2).split(' ', 1)[0]
            seen_response_line = True

    return method, url, headers, request_body, status_code, response_body


def build_curl(method, url, headers, body):
    if not method or not url:
        return ''
    parts = ['curl', '-X', method.upper()]
    for key, value in (headers or {}).items():
        if key.lower() in _NOISY_REQUEST_HEADERS:
            continue
        parts.append('-H')
        parts.append(_shell_quote(f'{key}: {value}'))
    parts.append(_shell_quote(url))
    if body:
        parts.append('-d')
        parts.append(_shell_quote(body))
    return ' '.join(parts)


def _iter_all_steps(scenario):
    """Yields a scenario's steps in execution order, inlining the steps of
    any called feature's scenarios at the point where the call happened
    (Karate runs `call`/`callSingle` synchronously right there) -- so a
    caller's own steps and everything a called feature did come out as one
    flat, correctly-ordered sequence, however many calls deep."""
    for step in scenario.get('steps', []):
        yield step
        if step.get('hasCallResults'):
            for call_result in step.get('callResults') or []:
                for called_scenario in call_result.get('scenarios', []):
                    yield from _iter_all_steps(called_scenario)


def extract_api_steps(scenario, warnings=None):
    """Returns one dict per HTTP call the scenario made (directly or via a
    called feature), in the order they actually ran. `warnings` (if given)
    collects human-readable notes about any 'method' step that couldn't be
    turned into a normal row -- appended to, not returned, so a caller
    walking many scenarios/features can share one list."""
    steps = []
    for step in _iter_all_steps(scenario):
        if step.get('keyword') != 'method':
            continue
        step_label = f"scenario '{scenario.get('name') or '(unnamed)'}', step '{step.get('text')}' (line {step.get('line', '?')})"
        if step.get('status') == 'skipped':
            # Karate marks a step explicitly 'skipped' when it never ran at
            # all (an earlier step in the scenario failed first) -- there's
            # nothing to reconstruct, and a phantom "Execute the CURL" row
            # for a call that never happened would be misleading, so it's
            # dropped rather than shown as a failure.
            #
            # Note: a missing/empty logSegments on its own is NOT used as
            # this signal -- some Karate configs/versions omit detailed
            # request/response logging for calls that did execute (this
            # affected non-GET methods disproportionately in practice,
            # since bodyless GETs are more likely to still produce a
            # parseable log), and treating "no logs" the same as "skipped"
            # caused those real executed calls to silently vanish instead
            # of showing up with the "could not reconstruct" placeholder.
            if warnings is not None:
                warnings.append(f"{step_label}: step was skipped (an earlier step in the scenario failed first) -- omitted from the output.")
            continue
        log_segments = step.get('logSegments') or []
        method, url, headers, req_body, status_code, resp_body = _split_request_response(log_segments)
        curl = build_curl(method, url, headers, req_body)
        if not curl and warnings is not None:
            warnings.append(f"{step_label}: logs were present but a curl command could not be reconstructed from them.")
        steps.append({
            'method': method or (step.get('text') or '').upper(),
            'url': url,
            'curl': curl or f"# Could not reconstruct the request for step: {step.get('text')}",
            'status_code': status_code,
            'response_body': resp_body,
        })
    return steps


def build_test_cases(feature, warnings=None):
    """Returns [{name, steps: [...]}] for every scenario in a parsed feature
    that made at least one HTTP call -- scenarios with no API calls aren't
    meaningful as an "API test case" so they're skipped."""
    feature_name = feature.get('name') or os.path.basename(feature.get('path', ''))
    cases = []
    for scenario in feature.get('scenarios', []):
        api_steps = extract_api_steps(scenario, warnings=warnings)
        if not api_steps:
            continue
        name = scenario.get('name') or feature_name
        cases.append({'name': name, 'steps': api_steps})
    return cases


def _format_result(status_code, response_body):
    return 'Response Code: {}\nResponse Body:\n{}'.format(
        status_code or 'N/A', response_body or '(no response body)'
    )


def write_excel(test_cases, excel_path, *, environment, pre_requisite, created_by, sprint):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Test Cases'
    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    wrap = Alignment(wrap_text=True, vertical='top')
    merged_alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    # Columns that are constant for a whole test case (not per API call) get
    # merged into a single spanning cell across that case's step rows,
    # rather than repeating the same value on every row.
    merged_columns = (1, 2, 3, 4, 5, 11, 12)  # S.No, Test Case Name/Description, Environment, Pre-Requisite, Created By, Sprint

    for case_number, case in enumerate(test_cases, start=1):
        start_row = ws.max_row + 1
        for step_number, step in enumerate(case['steps'], start=1):
            result_text = _format_result(step['status_code'], step['response_body'])
            is_first_step = step_number == 1
            ws.append([
                case_number if is_first_step else None,
                case['name'] if is_first_step else None,
                case['name'] if is_first_step else None,
                environment if is_first_step else None,
                pre_requisite if is_first_step else None,
                step_number, 'Execute the CURL', step['curl'], result_text, result_text,
                created_by if is_first_step else None,
                sprint if is_first_step else None,
            ])
            for cell in ws[ws.max_row]:
                cell.alignment = wrap

        end_row = ws.max_row
        if end_row > start_row:
            for col in merged_columns:
                ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)
                ws.cell(row=start_row, column=col).alignment = merged_alignment

    widths = [8, 30, 30, 14, 20, 8, 20, 60, 45, 45, 14, 12]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    parent = os.path.dirname(excel_path)
    if parent:
        os.makedirs(parent, exist_ok=True)
    wb.save(excel_path)


def generate(job):
    """Runs the full pipeline for a KarateTestCaseJob, writing progress/
    results onto `job` and saving it. Safe to call from a background thread."""
    try:
        files = find_report_files(job.reports_dir)
        warnings = []
        all_cases = []
        feature_count = 0

        for path in files:
            try:
                feature = parse_report_file(path)
            except (OSError, json.JSONDecodeError) as exc:
                warnings.append(f'{path}: {exc}')
                continue
            if feature is None:
                continue
            feature_count += 1
            step_warnings = []
            all_cases.extend(build_test_cases(feature, warnings=step_warnings))
            warnings.extend(f'{path}: {w}' for w in step_warnings)

        if feature_count == 0:
            job.status = job.STATUS_FAILED
            job.error = f"No Karate feature HTML reports found under '{job.reports_dir}'."
            job.warnings = warnings
        elif not all_cases:
            job.status = job.STATUS_FAILED
            job.error = 'Found Karate reports, but no scenario made any HTTP calls to document.'
            job.warnings = warnings
        else:
            write_excel(
                all_cases, job.excel_path,
                environment=job.environment, pre_requisite=job.pre_requisite,
                created_by=job.created_by, sprint=job.sprint,
            )
            job.feature_count = feature_count
            job.scenario_count = len(all_cases)
            job.step_count = sum(len(c['steps']) for c in all_cases)
            job.warnings = warnings
            job.status = job.STATUS_COMPLETED
    except Exception as exc:  # noqa: BLE001 -- background thread, must not raise
        job.status = job.STATUS_FAILED
        job.error = str(exc)

    job.completed_at = timezone.now()
    job.save()
