import tempfile
from unittest.mock import MagicMock, patch

from django.test import TestCase
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework.test import APIClient

from apitester.models import ImportedRequest, TestCase as ApiTestCase, TestRun
from karate_tests.runner import COLUMNS


class _SyncThread:
    """Stand-in for threading.Thread that runs its target immediately, in
    the calling thread, when .start() is called -- same pattern used by
    credential_tester's tests, so the background-execution code path can be
    exercised deterministically instead of racing a real thread against the
    test's own transaction teardown."""
    def __init__(self, target=None, args=(), daemon=None):
        self._target = target
        self._args = args

    def start(self):
        self._target(*self._args)


class StopTestRunViewTests(TestCase):
    def setUp(self):
        self.client = APIClient()

    def _import(self):
        response = self.client.post(
            '/api/import-curl/', {'curl': 'curl https://api.example.com/x'}, format='json'
        )
        return response.json()

    def test_stop_on_completed_run_returns_400(self):
        imported = self._import()
        run = TestRun.objects.create(
            imported_request_id=imported['id'], status=TestRun.STATUS_COMPLETED
        )
        response = self.client.post(f'/api/test-runs/{run.id}/stop/')
        self.assertEqual(response.status_code, 400)
        run.refresh_from_db()
        self.assertFalse(run.stop_requested)

    def test_stop_on_already_stopped_run_returns_400(self):
        imported = self._import()
        run = TestRun.objects.create(
            imported_request_id=imported['id'], status=TestRun.STATUS_STOPPED
        )
        response = self.client.post(f'/api/test-runs/{run.id}/stop/')
        self.assertEqual(response.status_code, 400)

    def test_stop_on_running_run_sets_flag_but_leaves_status_running(self):
        # Flipping the status is the background thread's job, once it next
        # checks the flag -- the view itself only requests the stop.
        imported = self._import()
        run = TestRun.objects.create(
            imported_request_id=imported['id'], status=TestRun.STATUS_RUNNING
        )
        response = self.client.post(f'/api/test-runs/{run.id}/stop/')
        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertTrue(body['stop_requested'])
        self.assertEqual(body['status'], 'running')
        run.refresh_from_db()
        self.assertTrue(run.stop_requested)
        self.assertEqual(run.status, TestRun.STATUS_RUNNING)


class BackgroundRunnerStopBehaviorTests(TestCase):
    def setUp(self):
        self.client = APIClient()

    @patch('apitester.views.threading.Thread', _SyncThread)
    @patch('apitester.test_executor.requests.request')
    def test_setting_stop_requested_mid_run_halts_remaining_cases(self, mock_request):
        response_stub = MagicMock(status_code=200, headers={}, text='{}')
        imported = self.client.post(
            '/api/import-curl/', {'curl': 'curl https://api.example.com/x'}, format='json'
        ).json()

        call_count = {'n': 0}

        def side_effect(method, url, **kwargs):
            call_count['n'] += 1
            if call_count['n'] == 1:
                # Simulate a user clicking Stop while the first case's
                # request is still in flight.
                TestRun.objects.filter(imported_request_id=imported['id']).update(stop_requested=True)
            return response_stub

        mock_request.side_effect = side_effect

        # 'http_method' generates several cases on top of the always-included
        # baseline, so there's at least one still-pending case for the
        # background runner to skip once it notices the stop flag.
        response = self.client.post(
            f'/api/imported-requests/{imported["id"]}/test-runs/',
            {'categories': ['http_method']},
            format='json',
        )
        self.assertEqual(response.status_code, 201)
        run_id = response.json()['id']

        run = TestRun.objects.get(pk=run_id)
        self.assertEqual(run.status, TestRun.STATUS_STOPPED)
        self.assertIsNotNone(run.completed_at)
        self.assertTrue(run.test_cases.count() > 1)
        # Only the first case (in flight when Stop was clicked) actually ran.
        self.assertEqual(call_count['n'], 1)
        executed = [tc for tc in run.test_cases.all() if tc.executed_at]
        pending = [tc for tc in run.test_cases.all() if not tc.executed_at]
        self.assertEqual(len(executed), 1)
        self.assertTrue(len(pending) >= 1)

    @patch('apitester.views.threading.Thread', _SyncThread)
    @patch('apitester.test_executor.requests.request')
    def test_run_completes_normally_when_never_stopped(self, mock_request):
        mock_request.return_value = MagicMock(status_code=200, headers={}, text='{}')
        imported = self.client.post(
            '/api/import-curl/', {'curl': 'curl https://api.example.com/x'}, format='json'
        ).json()

        response = self.client.post(
            f'/api/imported-requests/{imported["id"]}/test-runs/', {'categories': []}, format='json'
        )
        self.assertEqual(response.status_code, 201)
        run = TestRun.objects.get(pk=response.json()['id'])
        self.assertEqual(run.status, TestRun.STATUS_COMPLETED)
        self.assertFalse(run.stop_requested)


class ExportTestRunExcelViewTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.imported = ImportedRequest.objects.create(
            raw_curl='curl https://api.example.com/x', method='GET', url='https://api.example.com/x',
            headers={}, body=None, body_raw=None, is_json_body=False,
        )

    def _make_run(self, run_status):
        return TestRun.objects.create(imported_request=self.imported, status=run_status)

    def _add_executed_case(self, run, **overrides):
        defaults = dict(
            test_run=run, category='baseline', description='Baseline (unmodified request)',
            request_method='GET', request_url='https://api.example.com/x',
            request_headers={'Accept': 'application/json'}, body_mode='none',
            status_code=200, response_body='{"ok": true}', executed_at=timezone.now(),
        )
        defaults.update(overrides)
        return ApiTestCase.objects.create(**defaults)

    def test_export_on_running_run_returns_400(self):
        run = self._make_run(TestRun.STATUS_RUNNING)
        response = self.client.post(f'/api/test-runs/{run.id}/export-excel/', {'excel_path': '/tmp/x.xlsx'})
        self.assertEqual(response.status_code, 400)

    def test_export_with_no_executed_cases_returns_400(self):
        run = self._make_run(TestRun.STATUS_COMPLETED)
        response = self.client.post(f'/api/test-runs/{run.id}/export-excel/', {'excel_path': '/tmp/x.xlsx'})
        self.assertEqual(response.status_code, 400)
        self.assertIn('no executed test cases', response.json()['error'])

    def test_export_rejects_relative_excel_path(self):
        run = self._make_run(TestRun.STATUS_COMPLETED)
        self._add_executed_case(run)
        response = self.client.post(f'/api/test-runs/{run.id}/export-excel/', {'excel_path': 'relative.xlsx'})
        self.assertEqual(response.status_code, 400)

    def test_export_writes_workbook_with_karate_style_columns(self):
        run = self._make_run(TestRun.STATUS_STOPPED)
        self._add_executed_case(
            run, category='baseline', description='Baseline (unmodified request)',
            request_method='GET', status_code=200, response_body='{"ok": true}',
        )
        self._add_executed_case(
            run, category='body_field_null', description="Set field 'name' to null",
            request_method='POST', request_url='https://api.example.com/x',
            request_headers={'Content-Type': 'application/json'}, body_mode='json',
            request_body={'name': None}, status_code=None, response_body=None,
            error='Connection timed out',
        )
        # Never-executed case (e.g. left pending by a Stop) must not appear in the export.
        ApiTestCase.objects.create(
            test_run=run, category='http_method', description='Send request using PUT instead of GET',
            request_method='PUT', request_url='https://api.example.com/x',
            request_headers={}, body_mode='none',
        )

        with tempfile.TemporaryDirectory() as out_dir:
            excel_path = f'{out_dir}/api-test-cases.xlsx'
            response = self.client.post(f'/api/test-runs/{run.id}/export-excel/', {
                'excel_path': excel_path,
                'environment': 'QA',
                'pre_requisite': 'User is logged in',
                'created_by': 'Ada',
                'sprint': 'Sprint 1',
                'lob': 'Payments',
                'vertical': 'Retail',
                'feasible_for_automation': 'Yes',
                'test_case_applicability': 'Regression',
                'labels': 'smoke',
                'test_case_status': 'Active',
            })
            self.assertEqual(response.status_code, 200)
            body = response.json()
            self.assertEqual(body['excel_path'], excel_path)
            self.assertEqual(body['exported_case_count'], 2)

            wb = load_workbook(excel_path)
            ws = wb.active
            header = [c.value for c in ws[1]]
            self.assertEqual(header, COLUMNS)

            rows = list(ws.iter_rows(min_row=2, values_only=True))
            self.assertEqual(len(rows), 2)  # one row per exported case (each is a single-step case)

            first = rows[0]
            self.assertEqual(first[0], 1)  # S.No
            self.assertEqual(first[1], 'baseline: Baseline (unmodified request)')
            self.assertIn('curl -X GET', first[7])
            self.assertIn('Response Code: 200', first[8])
            self.assertEqual(first[3], 'QA')
            self.assertEqual(first[12], 'Payments')
            self.assertEqual(first[17], 'Active')

            second = rows[1]
            self.assertEqual(second[0], 2)
            self.assertIn('curl -X POST', second[7])
            # An errored case's error message is shown as the result, not a status code.
            self.assertIn('Connection timed out', second[8])
